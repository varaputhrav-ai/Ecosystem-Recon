"""
Ecosystem Purchase Recon Dashboard
Swiggy Instamart Finance — Month-End Reconciliation
"""

import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

# ─────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────

PACKAGING_CODES = [
    'SFP','SP9','SP3','SPM','STP','SK2','SAP','SP5','SPE','SLP',
    'SHP','SPN','SPP','SIP','SBP','SV1','SKP','PMP','PMM','SPC',
    'PIM','PMC','LKR','PMJ','PMK','SBQ'
]

AMOUNT_TOLERANCE = 1.0

FIELD_LABELS = {
    'inv':          'Invoice Number',
    'seller_gstin': 'Seller GSTIN',
    'buyer_gstin':  'Buyer GSTIN',
    'total':        'Invoice Total',
    'seller_name':  'Seller Name',
    'buyer_name':   'Buyer Name / Entity',
    'taxable':      'Taxable Value',
    'tax':          'Tax Amount',
}

REQUIRED_FIELDS = ['inv', 'seller_gstin', 'buyer_gstin', 'total']
OPTIONAL_FIELDS = ['seller_name', 'buyer_name', 'taxable', 'tax']

# Hints for smart default guessing (lowercase substrings to match against column names)
FIELD_HINTS = {
    'inv':          ['invoice_no', 'invoice no', 'bill number', 'inv_no'],
    'seller_gstin': ['seller_gstin', 'supplier_gstin', 'vendor gstin', 'supplier_gstn'],
    'buyer_gstin':  ['customer_gstin', 'entity_gstin', 'buyer_gstin', 'entity gstin', 'entity_gstn'],
    'total':        ['invoice_value', 'amount for reco', 'total_amt_with_tax', 'gross total',
                     'invoice total', 'sum of item total'],
    'seller_name':  ['seller_entity', 'new_supplier_name', 'vendor name'],
    'buyer_name':   ['buyer_entity', 'inbound entity'],
    'taxable':      ['taxable_value', 'total_amt_without_tax', 'taxable'],
    'tax':          ['tax_value', 'item_tax_value', 'sum of total tax', 'tax amount'],
}

SOURCE_TYPES = ['Sales', 'WMS / Purchase', 'Zoho Books']

COLORS = {
    'hdr_dark':   '1F3864',
    'hdr_mid':    '2F5597',
    'orange':     'FCE4D6',
    'blue_light': 'DDEBF7',
    'yellow':     'FFE699',
    'balanced':   'C6EFCE',
    'unbalanced': 'FFC7CE',
    'brs_row':    'D6E4F0',
}

# ─────────────────────────────────────────────
# DATA LOADING
# ─────────────────────────────────────────────

def detect_header_row(rows):
    for i, row in enumerate(rows):
        if len([c for c in row if c is not None and str(c).strip()]) >= 4:
            return i
    return 0

def get_sheet_names(file_obj):
    wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    names = wb.sheetnames
    wb.close()
    file_obj.seek(0)
    return names

def load_sheet(file_obj, sheet_name):
    wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(min_row=1, max_row=10, values_only=True))
    wb.close()
    file_obj.seek(0)
    hdr_idx = detect_header_row(rows)
    df = pd.read_excel(file_obj, sheet_name=sheet_name, header=hdr_idx, engine='openpyxl')
    df = df.dropna(how='all').dropna(axis=1, how='all')
    df.columns = [str(c).strip() for c in df.columns]
    df = df[[c for c in df.columns
             if not c.startswith('Unnamed') and c not in ('.', '', 'None')]]
    file_obj.seek(0)
    return df

def to_num(s):
    return pd.to_numeric(s, errors='coerce').fillna(0)

def is_packaging(inv_no):
    inv = str(inv_no).upper()
    return any(code in inv for code in PACKAGING_CODES)

def norm_amount(val):
    try:
        return round(float(val), 2)
    except Exception:
        return 0.0

# ─────────────────────────────────────────────
# SMART COLUMN GUESSER
# ─────────────────────────────────────────────

def guess_column(columns, field):
    """Return first column that matches any hint for this field."""
    hints = FIELD_HINTS.get(field, [])
    cols_lower = {c.lower(): c for c in columns}
    for hint in hints:
        for col_lower, col_orig in cols_lower.items():
            if hint in col_lower or col_lower in hint:
                return col_orig
    return None

# ─────────────────────────────────────────────
# PREPARE DF WITH USER COLUMN MAPPING
# ─────────────────────────────────────────────

def prepare_with_cm(df, cm, source_type):
    df = df.copy()
    df.attrs['_cm'] = cm
    df.attrs['_source'] = source_type

    inv_col = cm.get('inv')
    sg_col  = cm.get('seller_gstin')
    bg_col  = cm.get('buyer_gstin')

    df['_inv'] = (
        df[inv_col].astype(str).str.upper().str.strip()
        if inv_col and inv_col in df.columns else ''
    )

    if sg_col and sg_col in df.columns and bg_col and bg_col in df.columns:
        df['_gstin_pair'] = (
            df[sg_col].astype(str).str.strip() + '|' +
            df[bg_col].astype(str).str.strip()
        )
    else:
        df['_gstin_pair'] = ''

    # Zoho: auto-compute tax from IGST/CGST/SGST if user didn't map a tax column
    if source_type == 'Zoho Books' and (not cm.get('tax') or cm['tax'] not in df.columns):
        igst_col = next((c for c in df.columns if 'igst' in c.lower()), None)
        cgst_col = next((c for c in df.columns if 'cgst' in c.lower()), None)
        sgst_col = next((c for c in df.columns if 'sgst' in c.lower()), None)
        igst = to_num(df[igst_col]) if igst_col else pd.Series(0, index=df.index)
        cgst = to_num(df[cgst_col]) if cgst_col else pd.Series(0, index=df.index)
        sgst = to_num(df[sgst_col]) if sgst_col else pd.Series(0, index=df.index)
        computed = (igst + cgst + sgst).round(2)
        if computed.sum() > 0:
            df['_computed_tax'] = computed
            cm = dict(cm)
            cm['tax'] = '_computed_tax'
            df.attrs['_cm'] = cm

    # WMS: flag packaging invoices
    if source_type == 'WMS / Purchase':
        df['_is_pkg'] = df['_inv'].apply(is_packaging)

    return df

def get_cm(df):
    return df.attrs.get('_cm', {})

# ─────────────────────────────────────────────
# RECON CORE
# ─────────────────────────────────────────────

def run_recon(s1, s1_name, s2, s2_name):
    cm1 = get_cm(s1)
    cm2 = get_cm(s2)

    bad = {'', 'NAN', 'NONE', 'NAT'}
    s1_inv = set(s1['_inv']) - bad
    s2_inv = set(s2['_inv']) - bad
    exact  = s1_inv & s2_inv
    only1  = s1_inv - s2_inv
    only2  = s2_inv - s1_inv

    s1_m = s1[s1['_inv'].isin(exact)].drop_duplicates('_inv').copy()
    s2_m = s2[s2['_inv'].isin(exact)].drop_duplicates('_inv').copy()

    total_col1 = cm1.get('total', '')
    total_col2 = cm2.get('total', '')
    sg1 = cm1.get('seller_gstin', '')
    bg1 = cm1.get('buyer_gstin', '')
    sg2 = cm2.get('seller_gstin', '')
    bg2 = cm2.get('buyer_gstin', '')

    rename1 = {'_gstin_pair': '_m_gp_s1'}
    rename2 = {'_gstin_pair': '_m_gp_s2'}
    if total_col1 and total_col1 in s1_m.columns:
        rename1[total_col1] = '_m_total_s1'
    if total_col2 and total_col2 in s2_m.columns:
        rename2[total_col2] = '_m_total_s2'
    if sg1 and sg1 in s1_m.columns:
        rename1[sg1] = '_m_sg_s1'
    if bg1 and bg1 in s1_m.columns:
        rename1[bg1] = '_m_bg_s1'
    if sg2 and sg2 in s2_m.columns:
        rename2[sg2] = '_m_sg_s2'
    if bg2 and bg2 in s2_m.columns:
        rename2[bg2] = '_m_bg_s2'

    s1_m = s1_m.rename(columns=rename1)
    s2_m = s2_m.rename(columns=rename2)

    merged = s1_m.merge(s2_m, on='_inv', suffixes=('_s1x', '_s2x'))

    if '_m_total_s1' in merged.columns and '_m_total_s2' in merged.columns:
        merged['_amt_diff']    = (to_num(merged['_m_total_s1']) - to_num(merged['_m_total_s2'])).round(2)
        merged['_discrepancy'] = merged['_amt_diff'].abs() > AMOUNT_TOLERANCE
    else:
        merged['_amt_diff']    = 0.0
        merged['_discrepancy'] = False

    if '_m_sg_s1' in merged.columns and '_m_sg_s2' in merged.columns:
        merged['_gstin_diff'] = (
            merged['_m_sg_s1'].astype(str).str.strip() != merged['_m_sg_s2'].astype(str).str.strip()
        ) | (
            merged['_m_bg_s1'].astype(str).str.strip() != merged['_m_bg_s2'].astype(str).str.strip()
        )
    else:
        merged['_gstin_diff'] = False

    s1_only = s1[s1['_inv'].isin(only1)].copy()
    s2_only = s2[s2['_inv'].isin(only2)].copy()

    # Secondary match: GSTIN pair + amount → detect wrong invoice number in GRN
    if len(s1_only) > 0 and len(s2_only) > 0 and total_col2:
        s2_keys = {
            r['_gstin_pair'] + '|' + str(norm_amount(r.get(total_col2, 0)))
            for _, r in s2_only.iterrows()
        }
        def sec_cat(row):
            key = row['_gstin_pair'] + '|' + str(norm_amount(row.get(total_col1, 0)))
            return (
                'Invoice No. Mismatch in GRN (Human Error) — Secondary Match on GSTIN+Amount'
                if key in s2_keys else None
            )
        s1_only['_sec_match'] = s1_only.apply(sec_cat, axis=1)
    else:
        s1_only['_sec_match'] = None

    s1_only['_category'] = s1_only.apply(
        lambda r: categorize_s1(r, s1_name, s2_name), axis=1)
    s2_only['_category'] = s2_only.apply(
        lambda r: categorize_s2(r, s1_name, s2_name), axis=1)

    return merged, s1_only, s2_only


def categorize_s1(row, s1_name, s2_name):
    if row.get('_sec_match'):
        return row['_sec_match']
    if 'Sales' in s1_name:
        if 'WMS' in s2_name:  return 'Goods in Transit — GRN Not Done Yet'
        if 'Zoho' in s2_name: return 'Not Booked in Zoho Books — Provision Required'
    elif 'WMS' in s1_name:
        if is_packaging(str(row.get('_inv', ''))):
            return 'Packing Material Purchase (Not a Sale)'
        if 'Sales' in s2_name: return 'GRN Done — Sales Pending / Prior Month Transit'
        if 'Zoho'  in s2_name: return 'GRN Done — Not Booked in Zoho Books'
    elif 'Zoho' in s1_name:
        if 'Sales' in s2_name: return 'Prior Month Sales Booked in Current Month'
        if 'WMS'   in s2_name: return 'Booked in Zoho — GRN Pending'
    return 'To Be Investigated'


def categorize_s2(row, s1_name, s2_name):
    if 'WMS' in s2_name:
        if is_packaging(str(row.get('_inv', ''))):
            return 'Packing Material Purchase (Not a Sale)'
        return 'GRN Done — Sales Pending / Prior Month Transit'
    elif 'Sales' in s2_name: return 'Goods in Transit — GRN Not Done Yet'
    elif 'Zoho'  in s2_name: return 'Prior Month Entry Booked in Zoho'
    return 'To Be Investigated'


def build_brs(s1, s1_name, s2, s2_name, matched, s1_only, s2_only):
    cm1 = get_cm(s1)
    cm2 = get_cm(s2)
    tc1 = cm1.get('total', '')
    tc2 = cm2.get('total', '')

    s1_tot  = to_num(s1[tc1]).sum()         if tc1 and tc1 in s1.columns              else 0
    s2_tot  = to_num(s2[tc2]).sum()         if tc2 and tc2 in s2.columns              else 0
    s1o_tot = to_num(s1_only[tc1]).sum()    if tc1 and tc1 in s1_only.columns and len(s1_only) > 0 else 0
    s2o_tot = to_num(s2_only[tc2]).sum()    if tc2 and tc2 in s2_only.columns and len(s2_only) > 0 else 0
    m_diff  = (
        to_num(matched['_m_total_s2']).sum() - to_num(matched['_m_total_s1']).sum()
        if '_m_total_s1' in matched.columns and '_m_total_s2' in matched.columns and len(matched) > 0
        else 0
    )
    reconciled = s1_tot - s1o_tot + s2o_tot + m_diff
    diff       = round(reconciled - s2_tot, 2)
    disc_cnt   = int(matched['_discrepancy'].sum()) if len(matched) > 0 else 0

    rows = [
        ('',    f'{s1_name} Total',                       round(s1_tot,    2), ''),
        ('(-)', f'In {s1_name} only (not in {s2_name})', -round(s1o_tot,  2), f'{len(s1_only)} invoices'),
        ('(+)', f'In {s2_name} only (not in {s1_name})',  round(s2o_tot,  2), f'{len(s2_only)} invoices'),
        ('(±)', 'Value diff on matched invoices',          round(m_diff,   2), f'{disc_cnt} invoices with discrepancy'),
        ('=',   f'Reconciled {s2_name} Total',            round(reconciled,2), ''),
        ('',    f'Actual {s2_name} Total',                round(s2_tot,   2), ''),
        ('',    'Difference',                              diff,
                '✓ Balanced' if abs(diff) < AMOUNT_TOLERANCE else '⚠ Investigate'),
    ]
    return pd.DataFrame(rows, columns=['', 'Particulars', 'Amount (₹)', 'Notes'])

# ─────────────────────────────────────────────
# DISPLAY HELPERS
# ─────────────────────────────────────────────

def display_cols(cm, df):
    keys = ['inv', 'seller_name', 'buyer_name', 'seller_gstin', 'buyer_gstin', 'taxable', 'tax', 'total']
    cols = []
    for k in keys:
        v = cm.get(k)
        if v and v in df.columns and not v.startswith('_'):
            cols.append(v)
    return cols

def style_brs(brs_df):
    def row_style(row):
        label = str(row['Particulars'])
        diff  = row['Amount (₹)']
        if label == 'Difference':
            color = '#C6EFCE' if abs(float(diff or 0)) < AMOUNT_TOLERANCE else '#FFC7CE'
            return [f'background-color:{color};font-weight:bold'] * len(row)
        if 'Total' in label or 'Reconciled' in label:
            return ['background-color:#D6E4F0;font-weight:bold'] * len(row)
        return [''] * len(row)
    return brs_df.style.apply(row_style, axis=1).format({'Amount (₹)': '{:,.2f}'})

def render_table(df, cm, height=400):
    show = display_cols(cm, df)
    if '_category' in df.columns:
        show = show + ['_category']
    show = [c for c in show if c in df.columns]
    if not show:
        st.info('No columns to display.')
        return
    rename_map = {
        cm.get('inv'):          'Invoice No.',
        cm.get('seller_name'):  'Seller',
        cm.get('buyer_name'):   'Buyer / Entity',
        cm.get('seller_gstin'): 'Seller GSTIN',
        cm.get('buyer_gstin'):  'Buyer GSTIN',
        cm.get('taxable'):      'Taxable (₹)',
        cm.get('tax'):          'Tax (₹)',
        cm.get('total'):        'Invoice Total (₹)',
        '_category':            'Category / Reason',
    }
    display_df = df[show].rename(columns={k: v for k, v in rename_map.items() if k and k in show})
    st.dataframe(display_df, use_container_width=True, height=height, hide_index=True)

# ─────────────────────────────────────────────
# EXCEL EXPORT
# ─────────────────────────────────────────────

def apply_fill(ws, row, c1, c2, color):
    fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    for c in range(c1, c2 + 1):
        ws.cell(row=row, column=c).fill = fill

def write_hdr(ws, row, labels, color, font_color='FFFFFF'):
    fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    for c, lbl in enumerate(labels, 1):
        cell = ws.cell(row=row, column=c, value=lbl)
        cell.fill = fill
        cell.font = Font(color=font_color, bold=True)
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

def write_rows(ws, df, start, fill_color=None, num_cols=None):
    num_cols = num_cols or []
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid') if fill_color else None
    for ri, (_, row) in enumerate(df.iterrows()):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=start + ri, column=ci, value=val)
            if fill:
                cell.fill = fill
            if ci in num_cols and isinstance(val, (int, float)):
                cell.number_format = '#,##0.00'
    return start + len(df)

def autofit(ws):
    for col in ws.columns:
        w = max((len(str(c.value or '')) for c in col if c.value), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 3, 45)

def build_excel(results):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Summary sheet
    ws = wb.create_sheet('Summary', 0)
    ws.cell(1, 1, 'Ecosystem Purchase Recon — Monthly Summary').font = Font(
        bold=True, size=14, color=COLORS['hdr_dark'])
    ws.cell(2, 1, f'Generated: {datetime.now().strftime("%d-%b-%Y %H:%M")}').font = Font(
        italic=True, size=9)
    hdrs = ['Recon', 'S1 Total (₹)', 'S2 Total (₹)', 'S1 Only', 'S2 Only',
            'Discrepancies', 'Difference (₹)', 'Status']
    write_hdr(ws, 4, hdrs, COLORS['hdr_dark'])
    for i, (name, brs, s1_only, s2_only, matched, cm1, cm2) in enumerate(results, 5):
        diff = brs.iloc[6]['Amount (₹)']
        disc = int(matched['_discrepancy'].sum()) if len(matched) > 0 else 0
        vals = [
            name,
            brs.iloc[0]['Amount (₹)'], brs.iloc[5]['Amount (₹)'],
            len(s1_only), len(s2_only), disc, float(diff),
            '✓ Balanced' if abs(float(diff or 0)) < AMOUNT_TOLERANCE else f'⚠ ₹{diff:,.2f}',
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(i, c, v)
            if c in (2, 3, 7):
                cell.number_format = '#,##0.00'
        clr = COLORS['balanced'] if abs(float(diff or 0)) < AMOUNT_TOLERANCE else COLORS['unbalanced']
        apply_fill(ws, i, 8, 8, clr)
    autofit(ws)
    ws.freeze_panes = 'A5'

    # One sheet per recon combination
    for name, brs, s1_only, s2_only, matched, cm1, cm2 in results:
        s1n, s2n = name.split(' vs ')
        ws = wb.create_sheet(name[:31])
        r = 1

        ws.cell(r, 1, f'BRS: {name}').font = Font(bold=True, size=13, color=COLORS['hdr_dark'])
        ws.cell(r + 1, 1, f'Generated: {datetime.now().strftime("%d-%b-%Y %H:%M")}').font = Font(
            italic=True, size=9)
        r += 3

        ws.cell(r, 1, 'RECONCILIATION STATEMENT').font = Font(bold=True, size=11)
        r += 1
        write_hdr(ws, r, list(brs.columns), COLORS['hdr_dark'])
        r += 1
        for _, brow in brs.iterrows():
            for c, val in enumerate(brow, 1):
                cell = ws.cell(r, c, val)
                if c == 3 and isinstance(val, (int, float)):
                    cell.number_format = '#,##0.00'
            lbl = str(brow['Particulars'])
            if 'Total' in lbl or 'Reconciled' in lbl:
                apply_fill(ws, r, 1, 4, COLORS['brs_row'])
                ws.cell(r, 1).font = Font(bold=True)
                ws.cell(r, 3).font = Font(bold=True)
            if lbl == 'Difference':
                diff_v = brow['Amount (₹)']
                apply_fill(ws, r, 1, 4,
                           COLORS['balanced'] if abs(float(diff_v or 0)) < AMOUNT_TOLERANCE
                           else COLORS['unbalanced'])
                ws.cell(r, 1).font = Font(bold=True)
            r += 1
        r += 2

        if len(s1_only) > 0:
            tc1 = cm1.get('total', '')
            tot = to_num(s1_only[tc1]).sum() if tc1 and tc1 in s1_only.columns else 0
            ws.cell(r, 1, f'In {s1n} Only — {len(s1_only)} invoices | ₹{tot:,.2f}').font = Font(
                bold=True, size=11, color=COLORS['hdr_dark'])
            r += 1
            show = display_cols(cm1, s1_only) + (['_category'] if '_category' in s1_only.columns else [])
            write_hdr(ws, r, show, COLORS['hdr_mid'])
            r += 1
            num_p = [i + 1 for i, c in enumerate(show)
                     if c in (cm1.get('taxable'), cm1.get('tax'), cm1.get('total'))]
            r = write_rows(ws, s1_only[show], r, COLORS['orange'], num_p)
            r += 2

        if len(s2_only) > 0:
            tc2 = cm2.get('total', '')
            tot = to_num(s2_only[tc2]).sum() if tc2 and tc2 in s2_only.columns else 0
            ws.cell(r, 1, f'In {s2n} Only — {len(s2_only)} invoices | ₹{tot:,.2f}').font = Font(
                bold=True, size=11, color=COLORS['hdr_dark'])
            r += 1
            show = display_cols(cm2, s2_only) + (['_category'] if '_category' in s2_only.columns else [])
            write_hdr(ws, r, show, COLORS['hdr_mid'])
            r += 1
            num_p = [i + 1 for i, c in enumerate(show)
                     if c in (cm2.get('taxable'), cm2.get('tax'), cm2.get('total'))]
            r = write_rows(ws, s2_only[show], r, COLORS['blue_light'], num_p)
            r += 2

        if len(matched) > 0:
            disc = matched[matched['_discrepancy'] | matched['_gstin_diff']].copy()
            if len(disc) > 0:
                ws.cell(r, 1, f'Discrepancies — {len(disc)} invoices').font = Font(
                    bold=True, size=11, color='FF0000')
                r += 1
                disc_cols = [c for c in [
                    '_inv', '_m_sg_s1', '_m_sg_s2', '_m_bg_s1', '_m_bg_s2',
                    '_m_total_s1', '_m_total_s2', '_amt_diff', '_gstin_diff',
                ] if c in disc.columns]
                write_hdr(ws, r, disc_cols, COLORS['hdr_mid'])
                r += 1
                num_p = [i + 1 for i, c in enumerate(disc_cols)
                         if 'total' in c or 'diff' in c]
                r = write_rows(ws, disc[disc_cols], r, COLORS['yellow'], num_p)

        autofit(ws)
        ws.freeze_panes = 'A4'

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ─────────────────────────────────────────────
# COLUMN MAPPER WIDGET
# ─────────────────────────────────────────────

def column_mapper_ui(source_key, df, default_name='Source'):
    """Render sheet column mapper. Returns (src_name, src_type, cm)."""
    cols = list(df.columns)
    none_opt = '— Not mapped —'
    col_options_req = cols
    col_options_opt = [none_opt] + cols

    saved = st.session_state.get(f'mapping_{source_key}', {})

    def default_idx(field, options):
        saved_col = saved.get(field)
        guessed   = guess_column(cols, field)
        pick = saved_col if saved_col in options else (guessed if guessed in options else None)
        return options.index(pick) if pick in options else 0

    c_left, c_right = st.columns(2)
    with c_left:
        src_name = st.text_input(
            'Source label',
            value=saved.get('name', default_name),
            key=f'name_{source_key}',
            help='Give this source a short name, e.g. Sales, WMS, Zoho Books',
        )
    with c_right:
        src_type = st.selectbox(
            'Source type',
            SOURCE_TYPES,
            index=SOURCE_TYPES.index(saved['type']) if saved.get('type') in SOURCE_TYPES else 0,
            key=f'type_{source_key}',
            help='Used to auto-categorise exceptions (Transit, Packaging, etc.)',
        )

    st.markdown('**Required** — used for invoice matching')
    r1, r2, r3, r4 = st.columns(4)
    with r1:
        inv_col = st.selectbox(
            'Invoice Number ✱', col_options_req,
            index=default_idx('inv', col_options_req),
            key=f'inv_{source_key}',
        )
    with r2:
        total_col = st.selectbox(
            'Invoice Total ✱', col_options_req,
            index=default_idx('total', col_options_req),
            key=f'total_{source_key}',
        )
    with r3:
        sg_col = st.selectbox(
            'Seller GSTIN ✱', col_options_req,
            index=default_idx('seller_gstin', col_options_req),
            key=f'sg_{source_key}',
        )
    with r4:
        bg_col = st.selectbox(
            'Buyer GSTIN ✱', col_options_req,
            index=default_idx('buyer_gstin', col_options_req),
            key=f'bg_{source_key}',
        )

    st.markdown('**Optional** — for display only')
    o1, o2, o3, o4 = st.columns(4)
    with o1:
        sname_col = st.selectbox(
            'Seller Name', col_options_opt,
            index=default_idx('seller_name', col_options_opt),
            key=f'sname_{source_key}',
        )
    with o2:
        bname_col = st.selectbox(
            'Buyer / Entity', col_options_opt,
            index=default_idx('buyer_name', col_options_opt),
            key=f'bname_{source_key}',
        )
    with o3:
        tax_col = st.selectbox(
            'Tax Amount', col_options_opt,
            index=default_idx('tax', col_options_opt),
            key=f'tax_{source_key}',
        )
    with o4:
        taxable_col = st.selectbox(
            'Taxable Value', col_options_opt,
            index=default_idx('taxable', col_options_opt),
            key=f'taxable_{source_key}',
        )

    cm = {
        'inv':          inv_col,
        'total':        total_col,
        'seller_gstin': sg_col,
        'buyer_gstin':  bg_col,
        'seller_name':  sname_col   if sname_col   != none_opt else None,
        'buyer_name':   bname_col   if bname_col   != none_opt else None,
        'tax':          tax_col     if tax_col     != none_opt else None,
        'taxable':      taxable_col if taxable_col != none_opt else None,
    }
    return src_name, src_type, cm

# ─────────────────────────────────────────────
# MAIN APP
# ─────────────────────────────────────────────

def main():
    st.set_page_config(
        page_title='Ecosystem Purchase Recon',
        page_icon='📊',
        layout='wide',
        initial_sidebar_state='collapsed',
    )
    st.markdown("""
    <style>
    .block-container{padding-top:1.2rem;padding-bottom:1rem}
    .stDownloadButton>button{background:#1F3864;color:white;border-radius:6px;font-weight:600}
    .stTabs [data-baseweb="tab"]{font-weight:600;font-size:0.85rem}
    </style>""", unsafe_allow_html=True)

    st.title('📊 Ecosystem Purchase Recon')
    st.caption('Swiggy Instamart Finance | Month-End Reconciliation')
    st.divider()

    # ── STEP 1: Upload ──────────────────────────
    st.subheader('Step 1 — Upload Source Files')
    st.caption(
        '**Option A:** Upload one combined workbook (all 3 sources in different sheets).  '
        '**Option B:** Upload 2–3 separate files, one per source.'
    )
    c1, c2, c3 = st.columns(3)
    with c1:
        st.markdown('**File 1**')
        f1 = st.file_uploader('File 1', type=['xlsx', 'xls'], key='f1',
                               label_visibility='collapsed')
        if f1:
            st.success(f'✓ {f1.name}')
    with c2:
        st.markdown('**File 2** *(optional)*')
        f2 = st.file_uploader('File 2', type=['xlsx', 'xls'], key='f2',
                               label_visibility='collapsed')
        if f2:
            st.success(f'✓ {f2.name}')
    with c3:
        st.markdown('**File 3** *(optional)*')
        f3 = st.file_uploader('File 3', type=['xlsx', 'xls'], key='f3',
                               label_visibility='collapsed')
        if f3:
            st.success(f'✓ {f3.name}')

    files = [f for f in [f1, f2, f3] if f is not None]
    if not files:
        st.info('👆 Upload at least one file to begin.')
        return

    st.divider()

    # ── STEP 2: Configure each source ──────────
    st.subheader('Step 2 — Configure Sources')
    st.caption('Pick the sheet for each source and map the columns. Smart guesses are pre-filled.')

    DEFAULT_NAMES  = ['Sales', 'WMS', 'Zoho Books']
    DEFAULT_TYPES  = ['Sales', 'WMS / Purchase', 'Zoho Books']
    source_data    = {}

    # If only 1 file uploaded → treat it as combined workbook: show 3 source configs all from same file
    # If multiple files → one source config per file
    if len(files) == 1:
        combined_file = files[0]
        try:
            all_sheets = get_sheet_names(combined_file)
        except Exception as e:
            st.error(f'Cannot read file: {e}')
            return

        st.info(f'📄 **{combined_file.name}** — {len(all_sheets)} sheets detected. '
                f'Configure each source below by selecting the appropriate sheet.')

        for idx, src_key in enumerate(['S1', 'S2', 'S3']):
            default_label = DEFAULT_NAMES[idx]
            with st.expander(f'Source {idx+1} — {default_label}', expanded=True):
                saved_sheet = st.session_state.get(f'sheet_{src_key}')
                default_sheet_idx = (
                    all_sheets.index(saved_sheet) if saved_sheet in all_sheets
                    else min(idx, len(all_sheets) - 1)
                )
                chosen_sheet = st.selectbox(
                    'Sheet',
                    all_sheets,
                    index=default_sheet_idx,
                    key=f'sheet_sel_{src_key}',
                )
                st.session_state[f'sheet_{src_key}'] = chosen_sheet

                try:
                    combined_file.seek(0)
                    df_raw = load_sheet(combined_file, chosen_sheet)
                except Exception as e:
                    st.error(f'Could not load sheet "{chosen_sheet}": {e}')
                    continue

                st.caption(
                    f'{len(df_raw):,} rows × {len(df_raw.columns)} columns  |  '
                    f'Columns: {", ".join(df_raw.columns[:8])}'
                    + (' …' if len(df_raw.columns) > 8 else '')
                )

                src_name, src_type, cm = column_mapper_ui(src_key, df_raw, default_name=default_label)
                st.session_state[f'mapping_{src_key}'] = {'name': src_name, 'type': src_type, **cm}
                source_data[src_key] = (src_name, src_type, df_raw, cm)

    else:
        # Multiple files — one source per file
        for idx, fobj in enumerate(files):
            src_key       = f'S{idx+1}'
            default_label = DEFAULT_NAMES[idx] if idx < len(DEFAULT_NAMES) else f'Source {idx+1}'
            with st.expander(f'📄 {fobj.name}  ({default_label})', expanded=True):
                try:
                    sheets = get_sheet_names(fobj)
                except Exception as e:
                    st.error(f'Cannot read file: {e}')
                    continue

                saved_sheet       = st.session_state.get(f'sheet_{src_key}')
                default_sheet_idx = sheets.index(saved_sheet) if saved_sheet in sheets else 0
                chosen_sheet      = st.selectbox(
                    'Sheet', sheets,
                    index=default_sheet_idx,
                    key=f'sheet_sel_{src_key}',
                )
                st.session_state[f'sheet_{src_key}'] = chosen_sheet

                try:
                    df_raw = load_sheet(fobj, chosen_sheet)
                except Exception as e:
                    st.error(f'Could not load sheet "{chosen_sheet}": {e}')
                    continue

                st.caption(
                    f'{len(df_raw):,} rows × {len(df_raw.columns)} columns  |  '
                    f'Columns: {", ".join(df_raw.columns[:8])}'
                    + (' …' if len(df_raw.columns) > 8 else '')
                )

                src_name, src_type, cm = column_mapper_ui(src_key, df_raw, default_name=default_label)
                st.session_state[f'mapping_{src_key}'] = {'name': src_name, 'type': src_type, **cm}
                source_data[src_key] = (src_name, src_type, df_raw, cm)

    if len(source_data) < 2:
        st.warning('Configure at least 2 sources above to enable reconciliation.')
        return

    # ── STEP 3: Run ─────────────────────────────
    st.divider()
    if not st.button('▶  Run Reconciliation', type='primary', use_container_width=True):
        st.caption('Review column mappings above, then click Run.')
        return

    # Prepare dataframes with column maps
    prepared = {}
    with st.spinner('Preparing data…'):
        for src_key, (src_name, src_type, df_raw, cm) in source_data.items():
            df_prep = prepare_with_cm(df_raw, cm, src_type)
            prepared[src_key] = (src_name, src_type, df_prep)

    counts_str = '  |  '.join(f'{n}: {len(df):,} rows' for _, (n, _, df) in prepared.items())
    st.success(f'✓ {counts_str}')

    # All ordered combinations: S1 vs S2, S2 vs S1, S1 vs S3, S3 vs S1, S2 vs S3, S3 vs S2
    keys = list(prepared.keys())
    pairs = []
    for i in range(len(keys)):
        for j in range(len(keys)):
            if i != j:
                ki, kj = keys[i], keys[j]
                n1, _, d1 = prepared[ki]
                n2, _, d2 = prepared[kj]
                pairs.append((n1, n2, d1, d2))

    prog = st.progress(0)
    stxt = st.empty()
    results = []
    for idx, (s1n, s2n, s1, s2) in enumerate(pairs):
        stxt.text(f'Running {s1n} vs {s2n}…')
        matched, s1_only, s2_only = run_recon(s1, s1n, s2, s2n)
        brs = build_brs(s1, s1n, s2, s2n, matched, s1_only, s2_only)
        results.append((
            f'{s1n} vs {s2n}', brs, s1_only, s2_only, matched,
            get_cm(s1), get_cm(s2),
        ))
        prog.progress((idx + 1) / len(pairs))
    stxt.empty()
    prog.empty()

    # ── SUMMARY BANNER ──────────────────────────
    st.divider()
    st.subheader('📋 Recon Summary')
    mcols = st.columns(len(results))
    for col, (name, brs, s1_only, s2_only, matched, cm1, cm2) in zip(mcols, results):
        diff     = float(brs.iloc[6]['Amount (₹)'] or 0)
        balanced = abs(diff) < AMOUNT_TOLERANCE
        col.metric(
            label=name,
            value='✅ Balanced' if balanced else '⚠️ Diff',
            delta=(f'₹{diff:,.0f}' if not balanced else f'{len(s1_only) + len(s2_only)} items'),
            delta_color='normal' if balanced else 'inverse',
        )

    t1, t2, t3, t4 = st.columns(4)
    t1.metric('Balanced', f'{sum(1 for r in results if abs(float(r[1].iloc[6]["Amount (₹)"] or 0)) < AMOUNT_TOLERANCE)} / {len(results)}')
    t2.metric('Total S1-Only Items', sum(len(r[2]) for r in results))
    t3.metric('Total S2-Only Items', sum(len(r[3]) for r in results))
    t4.metric('Total Discrepancies', sum(
        int(r[4]['_discrepancy'].sum()) if len(r[4]) > 0 else 0 for r in results))

    st.divider()

    # ── FILTERS ─────────────────────────────────
    st.subheader('🔍 Filters')
    fc1, fc2 = st.columns([2, 3])
    with fc1:
        inv_search = st.text_input('Search Invoice Number', placeholder='Type invoice no…')
    with fc2:
        entity_filter = []
        first_key = keys[0]
        _, _, first_df = prepared[first_key]
        first_cm = get_cm(first_df)
        buyer_col = first_cm.get('buyer_name')
        if buyer_col and buyer_col in first_df.columns:
            entities = sorted(first_df[buyer_col].dropna().astype(str).unique().tolist())
            entity_filter = st.multiselect('Filter by Buyer / Entity', entities, placeholder='All entities')

    st.divider()

    # ── RECON TABS ───────────────────────────────
    tabs = st.tabs([r[0] for r in results])

    for tab, (name, brs, s1_only, s2_only, matched, cm1, cm2) in zip(tabs, results):
        s1n, s2n = name.split(' vs ')
        with tab:
            st.markdown(f'#### {name}')
            diff_val = float(brs.iloc[6]['Amount (₹)'] or 0)
            if abs(diff_val) < AMOUNT_TOLERANCE:
                st.success('✅ BRS balanced — sources reconcile.')
            else:
                st.error(f'⚠️ Unreconciled difference: ₹{diff_val:,.2f}')
            st.dataframe(style_brs(brs), use_container_width=True, hide_index=True)
            st.markdown('---')

            def apply_filters(df, cm):
                out = df.copy()
                if inv_search:
                    ic = cm.get('inv')
                    if ic and ic in out.columns:
                        out = out[out[ic].astype(str).str.upper().str.contains(
                            inv_search.upper(), na=False)]
                if entity_filter:
                    bc = cm.get('buyer_name')
                    if bc and bc in out.columns:
                        out = out[out[bc].astype(str).isin(entity_filter)]
                return out

            s1f        = apply_filters(s1_only, cm1)
            s2f        = apply_filters(s2_only, cm2)
            disc_count = int(matched['_discrepancy'].sum()) if len(matched) > 0 else 0

            sub1, sub2, sub3 = st.tabs([
                f'In {s1n} Only ({len(s1f):,})',
                f'In {s2n} Only ({len(s2f):,})',
                f'Discrepancies ({disc_count:,})',
            ])

            with sub1:
                if len(s1f) == 0:
                    st.success('All reconciled.' if not inv_search and not entity_filter
                               else 'No results match filter.')
                else:
                    cats     = ['All'] + sorted(s1f['_category'].dropna().unique().tolist())
                    sel      = st.selectbox('Filter by Category', cats, key=f's1cat_{name}')
                    df_show  = s1f if sel == 'All' else s1f[s1f['_category'] == sel]
                    cat_cnts = s1f['_category'].value_counts()
                    cc = st.columns(min(len(cat_cnts), 4))
                    for idx2, (cat, cnt) in enumerate(cat_cnts.items()):
                        cc[idx2 % 4].metric(cat[:40], cnt)
                    tc  = cm1.get('total')
                    tot = to_num(df_show[tc]).sum() if tc and tc in df_show.columns else 0
                    st.caption(f'{len(df_show):,} invoices | ₹{tot:,.2f}')
                    render_table(df_show, cm1,
                                 height=min(400, max(150, len(df_show) * 35 + 50)))

            with sub2:
                if len(s2f) == 0:
                    st.success('All reconciled.' if not inv_search and not entity_filter
                               else 'No results match filter.')
                else:
                    cats     = ['All'] + sorted(s2f['_category'].dropna().unique().tolist())
                    sel      = st.selectbox('Filter by Category', cats, key=f's2cat_{name}')
                    df_show  = s2f if sel == 'All' else s2f[s2f['_category'] == sel]
                    cat_cnts = s2f['_category'].value_counts()
                    cc = st.columns(min(len(cat_cnts), 4))
                    for idx2, (cat, cnt) in enumerate(cat_cnts.items()):
                        cc[idx2 % 4].metric(cat[:40], cnt)
                    tc  = cm2.get('total')
                    tot = to_num(df_show[tc]).sum() if tc and tc in df_show.columns else 0
                    st.caption(f'{len(df_show):,} invoices | ₹{tot:,.2f}')
                    render_table(df_show, cm2,
                                 height=min(400, max(150, len(df_show) * 35 + 50)))

            with sub3:
                if disc_count == 0:
                    st.success('No discrepancies — all matched invoices are consistent.')
                else:
                    disc = matched[matched['_discrepancy'] | matched['_gstin_diff']].copy()
                    if inv_search:
                        disc = disc[disc['_inv'].str.upper().str.contains(
                            inv_search.upper(), na=False)]
                    disc_cols = [c for c in [
                        '_inv', '_m_sg_s1', '_m_sg_s2', '_m_bg_s1', '_m_bg_s2',
                        '_m_total_s1', '_m_total_s2', '_amt_diff', '_gstin_diff',
                    ] if c in disc.columns]
                    rename = {
                        '_inv':         'Invoice No.',
                        '_m_sg_s1':     f'Seller GSTIN ({s1n})',
                        '_m_sg_s2':     f'Seller GSTIN ({s2n})',
                        '_m_bg_s1':     f'Buyer GSTIN ({s1n})',
                        '_m_bg_s2':     f'Buyer GSTIN ({s2n})',
                        '_m_total_s1':  f'Total ({s1n}) ₹',
                        '_m_total_s2':  f'Total ({s2n}) ₹',
                        '_amt_diff':    'Diff (₹)',
                        '_gstin_diff':  'GSTIN Mismatch',
                    }
                    st.caption(f'{len(disc):,} discrepancies')
                    st.dataframe(
                        disc[disc_cols].rename(columns=rename),
                        use_container_width=True, height=400, hide_index=True,
                    )

    # ── DOWNLOAD ─────────────────────────────────
    st.divider()
    with st.spinner('Building Excel…'):
        excel_buf = build_excel(results)
    st.download_button(
        label='📥 Download Recon Output (.xlsx)',
        data=excel_buf,
        file_name=f'Ecosystem_Recon_{datetime.now().strftime("%b_%Y")}.xlsx',
        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        use_container_width=True,
    )
    st.caption('Swiggy Instamart Finance | Ecosystem Purchase Recon Tool')


if __name__ == '__main__':
    main()
